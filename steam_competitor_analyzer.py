#!/usr/bin/env python3
import os, time, json, re, html, sys
from datetime import datetime
from pathlib import Path

import requests
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ══════════════════════════════════════════════════════════
#  ① 설정 로드 (config.yaml)
# ══════════════════════════════════════════════════════════

def _load_config():
    try:
        import yaml
    except ImportError:
        print("[오류] pyyaml이 설치되지 않았습니다: pip install pyyaml")
        sys.exit(1)
    cfg_path = Path(__file__).parent / "config.yaml"
    if not cfg_path.exists():
        print(f"[오류] 설정 파일 없음: {cfg_path}")
        print("  config.example.yaml 을 config.yaml 로 복사 후 설정을 입력하세요.")
        sys.exit(1)
    with open(cfg_path, encoding="utf-8") as f:
        return yaml.safe_load(f)

_cfg = _load_config()

APP_IDS              = _cfg.get("app_ids", [])
_chrome              = _cfg.get("chrome", {})
CHROME_USER_DATA_DIR = os.environ.get(
    "CHROME_USER_DATA_DIR",
    _chrome.get("user_data_dir", str(Path.home() / "AppData/Local/Google/Chrome/User Data")),
)
CHROME_PROFILE       = _chrome.get("profile", "Default")
ANTHROPIC_API_KEY    = os.environ.get("ANTHROPIC_API_KEY") or _cfg.get("anthropic_api_key", "")
MY_GAME_CONTEXT      = _cfg.get("game_context", "").strip()
OUTPUT_PATH          = (
    f"{_cfg.get('output_prefix', 'steam_competitor_analysis')}"
    f"_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
)
STEAMDB_WAIT_SECONDS = _cfg.get("steamdb_wait_seconds", 12)

# ══════════════════════════════════════════════════════════
#  ② Steam / SteamSpy API (requests 기반, 빠름)
# ══════════════════════════════════════════════════════════

OWNERS_MIDPOINT = {
    "0 .. 20,000": 10_000,
    "20,000 .. 50,000": 35_000,
    "50,000 .. 100,000": 75_000,
    "100,000 .. 200,000": 150_000,
    "200,000 .. 500,000": 350_000,
    "500,000 .. 1,000,000": 750_000,
    "1,000,000 .. 2,000,000": 1_500_000,
    "2,000,000 .. 5,000,000": 3_500_000,
    "5,000,000 .. 10,000,000": 7_500_000,
    "10,000,000 .. 20,000,000": 15_000_000,
    "20,000,000 .. 50,000,000": 35_000_000,
}

def strip_html(t):
    return html.unescape(re.sub(r"<[^>]+>", "", t or "")).strip()

def estimate_revenue(owners_str, price_cents):
    if not owners_str or not price_cents:
        return "N/A"
    nums = [int(n.replace(",", "")) for n in re.findall(r"[\d,]+", owners_str)]
    mid = OWNERS_MIDPOINT.get(owners_str, sum(nums) / len(nums) if nums else 0)
    est = mid * (price_cents / 100) * 0.7 * 0.5
    if est >= 1_000_000: return f"~${est/1_000_000:.1f}M"
    if est >= 1_000:     return f"~${est/1_000:.0f}K"
    return f"~${est:.0f}"

def format_tags(d, n=12):
    if not d: return ""
    return ", ".join(t for t, _ in sorted(d.items(), key=lambda x: -x[1])[:n])

def collect_steam_data(appid):
    """Steam Store API + SteamSpy로 기본 데이터 수집."""
    try:
        r = requests.get(
            f"https://store.steampowered.com/api/appdetails?appids={appid}&cc=us&l=english",
            timeout=15)
        d = r.json().get(str(appid), {})
        steam = d.get("data", {}) if d.get("success") else {}
    except Exception as e:
        print(f"    [Steam 오류] {e}"); steam = {}
    time.sleep(1.5)

    try:
        spy = requests.get(
            f"https://steamspy.com/api.php?request=appdetails&appid={appid}",
            timeout=15).json()
    except Exception as e:
        print(f"    [SteamSpy 오류] {e}"); spy = {}
    time.sleep(1.2)

    name        = steam.get("name") or spy.get("name") or f"App {appid}"
    devs        = ", ".join(steam.get("developers", [spy.get("developer", "")])) or ""
    pubs        = ", ".join(steam.get("publishers", [spy.get("publisher", "")])) or ""
    rel         = steam.get("release_date", {})
    release     = rel.get("date", "") if isinstance(rel, dict) else ""
    short_desc  = strip_html(steam.get("short_description", ""))
    pd          = steam.get("price_overview", {})
    price_cents = pd.get("initial", 0) if pd else 0
    price_str   = f"${price_cents/100:.2f}" if price_cents else "Free / N/A"
    genres      = ", ".join(g.get("description","") for g in steam.get("genres",[]))
    plat        = steam.get("platforms", {})
    platforms   = ", ".join(p.capitalize() for p in ("windows","mac","linux") if plat.get(p))
    meta        = steam.get("metacritic", {})
    meta_score  = str(meta.get("score","")) if isinstance(meta,dict) and meta else ""
    pos, neg    = spy.get("positive",0), spy.get("negative",0)
    total_rev   = pos + neg
    pos_pct     = f"{pos/total_rev*100:.1f}%" if total_rev else ""
    owners_str  = spy.get("owners","")
    avg_pt      = spy.get("average_forever", 0)
    avg_pt_str  = f"{avg_pt//60}h {avg_pt%60}m" if avg_pt else ""
    tags_str    = format_tags(spy.get("tags",{}))
    est_rev     = estimate_revenue(owners_str, price_cents)

    return {
        "appid": appid,
        "name": name,
        "steam_url": f"https://store.steampowered.com/app/{appid}/",
        "developer": devs,
        "publisher": pubs,
        "release_date": release,
        "total_reviews": total_rev or "",
        "positive_pct": pos_pct,
        "owners_range": owners_str,
        "estimated_rev": est_rev,
        "price": price_str,
        "avg_playtime": avg_pt_str,
        "genres": genres,
        "platforms": platforms,
        "metacritic": meta_score,
        "tags": tags_str,
        "short_desc": short_desc,
        # Playwright가 채워줌
        "followers": "수집 중...",
        "wishlist": "수집 중...",
        # Claude가 채워줌
        "core_genre": "",
        "visual_tone": "",
        "differentiator": "",
        "notes": "",
    }

# ══════════════════════════════════════════════════════════
#  ③ Playwright — SteamDB Extension 크롤링
# ══════════════════════════════════════════════════════════

# SteamDB Extension이 주입하는 DOM 요소를 찾는 CSS 셀렉터 목록
# Extension 업데이트에 따라 셀렉터가 바뀔 수 있으므로,
# 동작 안 하면 Chrome DevTools > Inspect 으로 직접 확인 후 수정하세요.
FOLLOWER_SELECTORS = [
    # SteamDB Extension이 주입하는 follower 수 위치 (버전마다 다를 수 있음)
    ".app-stat .num[data-tooltip*='follower']",
    "[data-steamdb-stat='followers'] .num",
    ".steamdb_stats .followers",
    # 직접 텍스트가 포함된 dt/dd 구조
    "dt:contains('Followers') + dd",
]

WISHLIST_SELECTORS = [
    ".app-stat .num[data-tooltip*='wishlist']",
    "[data-steamdb-stat='wishlists'] .num",
    ".steamdb_stats .wishlists",
    "dt:contains('Wishlists') + dd",
]

def parse_number(text: str) -> str:
    """'1,234,567' → '1,234,567' 형태로 정제."""
    if not text:
        return ""
    text = text.strip().replace("\xa0", "")
    # 이미 포맷된 숫자면 그대로 반환
    if re.match(r"^[\d,]+$", text.replace(" ", "")):
        return text
    # 숫자만 추출
    nums = re.findall(r"\d+", text.replace(",", "").replace(".", ""))
    return nums[0] if nums else text

async def scrape_steamdb_with_extension(page, appid: int) -> dict:
    """
    Playwright Page 객체로 스팀 페이지를 열고
    SteamDB Extension이 주입한 데이터를 읽습니다.
    """
    url = f"https://store.steampowered.com/app/{appid}/"
    result = {"followers": "N/A (셀렉터 미일치)", "wishlist": "N/A (셀렉터 미일치)"}

    try:
        await page.goto(url, wait_until="domcontentloaded", timeout=30000)

        # 나이 확인 팝업 자동 처리
        try:
            btn = page.locator("a#age_gate_btn_continue, #ageYear")
            if await btn.count() > 0:
                # 나이 선택 후 계속 클릭
                await page.select_option("#ageYear", "1990")
                await page.click("a#age_gate_btn_continue")
                await page.wait_for_load_state("domcontentloaded")
        except Exception:
            pass

        # Extension이 데이터를 주입할 때까지 대기
        # 방법 1: 특정 셀렉터가 나타날 때까지 최대 N초 대기
        injected = False
        deadline = time.time() + STEAMDB_WAIT_SECONDS
        while time.time() < deadline and not injected:
            for sel in FOLLOWER_SELECTORS + WISHLIST_SELECTORS:
                try:
                    count = await page.locator(sel).count()
                    if count > 0:
                        injected = True
                        break
                except Exception:
                    pass
            if not injected:
                await page.wait_for_timeout(500)

        if not injected:
            # Extension 데이터 없음 → 페이지 소스에서 직접 파싱 시도
            print(f"    [경고] SteamDB 데이터 주입 확인 안 됨 (appid={appid})")

        # Followers 읽기
        for sel in FOLLOWER_SELECTORS:
            try:
                el = page.locator(sel).first
                if await el.count() > 0:
                    txt = await el.inner_text()
                    val = parse_number(txt)
                    if val:
                        result["followers"] = val
                        break
            except Exception:
                continue

        # Wishlist 읽기
        for sel in WISHLIST_SELECTORS:
            try:
                el = page.locator(sel).first
                if await el.count() > 0:
                    txt = await el.inner_text()
                    val = parse_number(txt)
                    if val:
                        result["wishlist"] = val
                        break
            except Exception:
                continue

        # 셀렉터로 못 찾은 경우 — 페이지 전체 텍스트에서 패턴 검색 (폴백)
        if result["followers"] == "N/A (셀렉터 미일치)":
            content = await page.content()
            # SteamDB Extension이 주입한 텍스트 패턴 탐색
            m = re.search(r"[Ff]ollowers?[^\d]*?([\d,]+)", content)
            if m:
                result["followers"] = m.group(1)

        if result["wishlist"] == "N/A (셀렉터 미일치)":
            if 'content' not in dir():
                content = await page.content()
            m = re.search(r"[Ww]ishlists?[^\d]*?([\d,]+)", content)
            if m:
                result["wishlist"] = m.group(1)

    except Exception as e:
        print(f"    [Playwright 오류] appid={appid}: {e}")
        result = {"followers": f"오류: {str(e)[:40]}", "wishlist": "오류"}

    return result


def run_playwright_scraping(all_data: list) -> list:
    """
    Playwright를 사용해 모든 게임의 Followers/Wishlist를 수집합니다.
    동기 래퍼 — 내부적으로 asyncio를 사용합니다.
    """
    try:
        from playwright.sync_api import sync_playwright
    except ImportError:
        print("\n[오류] playwright가 설치되지 않았습니다.")
        print("  pip install playwright")
        print("  playwright install chromium")
        for g in all_data:
            g["followers"] = "playwright 미설치"
            g["wishlist"]  = "playwright 미설치"
        return all_data

    print(f"\n🌐 Playwright 시작 — Chrome 프로파일: {CHROME_USER_DATA_DIR}/{CHROME_PROFILE}")
    print("   ⚠  Chrome이 완전히 닫혀 있어야 합니다!")

    with sync_playwright() as p:
        # 실제 Chrome(Chromium) 실행, Extension이 설치된 프로파일 사용
        context = p.chromium.launch_persistent_context(
            user_data_dir=CHROME_USER_DATA_DIR,
            channel="chrome",         # 시스템 Chrome 사용 (chromium으로 바꾸면 Extension 없음)
            headless=False,           # Extension 활성화를 위해 반드시 False
            args=[
                f"--profile-directory={CHROME_PROFILE}",
                "--disable-blink-features=AutomationControlled",
                "--no-first-run",
                "--no-default-browser-check",
            ],
            viewport={"width": 1280, "height": 900},
            locale="en-US",
        )

        page = context.new_page()

        for i, game in enumerate(all_data, 1):
            appid = game["appid"]
            print(f"  [{i}/{len(all_data)}] {game['name']} (appid={appid}) — 브라우저 크롤링...")

            # sync_playwright에서는 동기 API 사용
            result = _scrape_sync(page, appid)
            game["followers"] = result.get("followers", "N/A")
            game["wishlist"]  = result.get("wishlist", "N/A")
            print(f"    Followers: {game['followers']}  |  Wishlist: {game['wishlist']}")

            # 페이지 간 딜레이 (Extension 로딩 시간 + rate limit)
            time.sleep(3)

        context.close()

    return all_data


def _scrape_sync(page, appid: int) -> dict:
    """sync_playwright용 동기 버전."""
    from playwright.sync_api import TimeoutError as PWTimeout

    url = f"https://store.steampowered.com/app/{appid}/"
    result = {"followers": "N/A", "wishlist": "N/A"}

    try:
        page.goto(url, wait_until="domcontentloaded", timeout=30000)

        # 나이 확인 팝업
        try:
            if page.locator("#ageYear").count() > 0:
                page.select_option("#ageYear", "1990")
                page.click("a#age_gate_btn_continue")
                page.wait_for_load_state("domcontentloaded")
        except Exception:
            pass

        # Extension 주입 대기
        injected = False
        deadline = time.time() + STEAMDB_WAIT_SECONDS
        while time.time() < deadline and not injected:
            for sel in FOLLOWER_SELECTORS + WISHLIST_SELECTORS:
                try:
                    if page.locator(sel).count() > 0:
                        injected = True
                        break
                except Exception:
                    pass
            if not injected:
                time.sleep(0.5)

        # Followers
        for sel in FOLLOWER_SELECTORS:
            try:
                el = page.locator(sel).first
                if el.count() > 0:
                    val = parse_number(el.inner_text())
                    if val:
                        result["followers"] = val
                        break
            except Exception:
                continue

        # Wishlist
        for sel in WISHLIST_SELECTORS:
            try:
                el = page.locator(sel).first
                if el.count() > 0:
                    val = parse_number(el.inner_text())
                    if val:
                        result["wishlist"] = val
                        break
            except Exception:
                continue

        # 폴백: 전체 소스에서 패턴 검색
        if result["followers"] == "N/A" or result["wishlist"] == "N/A":
            content = page.content()
            if result["followers"] == "N/A":
                m = re.search(r"[Ff]ollowers?[^\d]{0,20}([\d,]+)", content)
                if m: result["followers"] = m.group(1)
            if result["wishlist"] == "N/A":
                m = re.search(r"[Ww]ishlists?[^\d]{0,20}([\d,]+)", content)
                if m: result["wishlist"] = m.group(1)

    except PWTimeout:
        print(f"    [타임아웃] appid={appid}")
        result = {"followers": "타임아웃", "wishlist": "타임아웃"}
    except Exception as e:
        print(f"    [오류] appid={appid}: {e}")
        result = {"followers": f"오류", "wishlist": "오류"}

    return result

# ══════════════════════════════════════════════════════════
#  ④ Claude API — 메타 분석 (v2와 동일)
# ══════════════════════════════════════════════════════════

ANALYSIS_SYSTEM = """
당신은 스팀 인디 게임 마케팅 전문 애널리스트입니다.
게임 데이터를 분석해 마케팅 인사이트를 JSON으로 반환합니다.
반드시 아래 JSON 형식만 출력하고, 다른 텍스트는 절대 포함하지 마세요.

{
  "core_genre": "장르를 마케팅 관점에서 2~4단어로",
  "visual_tone": "비주얼 스타일·색감·아트 방향을 2~3문장으로",
  "differentiator": "경쟁 게임 대비 핵심 차별점을 2~3문장으로",
  "notes": "우리 게임 마케팅에 참고할 포인트 1문장"
}
"""

def analyze_with_claude(game: dict) -> dict:
    api_key = ANTHROPIC_API_KEY
    if not api_key or api_key.startswith("여기에"):
        return {"core_genre":"API키미설정","visual_tone":"","differentiator":"","notes":""}

    user_content = f"""
아래 스팀 게임을 분석해주세요.

== 분석 대상 ==
이름: {game['name']}
장르: {game['genres']}
태그: {game['tags']}
설명: {game['short_desc']}
출시일: {game['release_date']}  |  가격: {game['price']}
리뷰: {game['total_reviews']}개 ({game['positive_pct']} 긍정)  |  Followers: {game['followers']}

== 마케팅 컨텍스트 ==
{MY_GAME_CONTEXT}

JSON만 반환하세요.
"""
    try:
        r = requests.post(
            "https://api.anthropic.com/v1/messages",
            headers={"x-api-key": api_key, "anthropic-version": "2023-06-01",
                     "content-type": "application/json"},
            json={"model": "claude-sonnet-4-20250514", "max_tokens": 1000,
                  "system": ANALYSIS_SYSTEM,
                  "messages": [{"role": "user", "content": user_content}]},
            timeout=30,
        )
        r.raise_for_status()
        raw = re.sub(r"^```[a-z]*\n?|\n?```$", "", r.json()["content"][0]["text"].strip())
        res = json.loads(raw)
        return {k: res.get(k, "") for k in ("core_genre","visual_tone","differentiator","notes")}
    except Exception as e:
        print(f"    [Claude 오류] {e}")
        return {"core_genre": "분석오류", "visual_tone": str(e)[:60], "differentiator": "", "notes": ""}

# ══════════════════════════════════════════════════════════
#  ⑤ Excel 빌드 (v2와 동일한 스타일)
# ══════════════════════════════════════════════════════════

C = {"bg":"1C2030","s2":"1A3A6A","s3":"1A4A2A","s4":"3A2A5A","s5":"5A2A1A",
     "green":"00C896","row_a":"F8FAFF","row_b":"EEF3FF",
     "ai_a":"FDF8FF","ai_b":"F3EEFF","manual":"FFF3CD","border":"BCC8D8"}

def _fill(c): return PatternFill("solid", fgColor=c)
def _bdr():
    s = Side(style="thin", color=C["border"])
    return Border(left=s, right=s, top=s, bottom=s)

COLS = [
    ("No.",               None,           4.5, "center", "bg"),
    ("게임 이름",          "name",        24,  "left",   "bg"),
    ("Steam 링크",         "steam_url",   32,  "left",   "bg"),
    ("개발사",             "developer",   18,  "left",   "s2"),
    ("퍼블리셔",           "publisher",   18,  "left",   "s2"),
    ("출시일",             "release_date",13,  "center", "s2"),
    ("리뷰 수",            "total_reviews",10, "center", "s3"),
    ("긍정 비율",          "positive_pct",10,  "center", "s3"),
    ("Follower 수\n(자동)","followers",   16,  "center", "s3"),
    ("Wishlist 수\n(자동)","wishlist",    16,  "center", "s3"),
    ("소유자 수 범위",     "owners_range",24,  "left",   "s3"),
    ("예상 매출",          "estimated_rev",14, "center", "s3"),
    ("가격",               "price",       10,  "center", "s3"),
    ("평균 플레이타임",    "avg_playtime",14,  "center", "s3"),
    ("태그 (상위 12개)",  "tags",         46,  "left",   "s4"),
    ("Short Description", "short_desc",   52,  "left",   "s4"),
    ("Steam 장르",         "genres",      22,  "left",   "s4"),
    ("플랫폼",             "platforms",   14,  "left",   "s4"),
    ("메타크리틱",         "metacritic",  10,  "center", "s4"),
    ("핵심 장르 (AI)",     "core_genre",  28,  "left",   "s5"),
    ("비주얼/톤 (AI)",     "visual_tone", 40,  "left",   "s5"),
    ("차별점 (AI)",        "differentiator",40,"left",  "s5"),
    ("마케팅 참고 (AI)",   "notes",       30,  "left",   "s5"),
]

SEC_LABEL = {
    "bg": "기본 정보",
    "s2": "개발·출시",
    "s3": "수치 데이터  ★ Followers/Wishlist 자동 수집",
    "s4": "게임 콘텐츠",
    "s5": "🤖 AI 분석 (Claude)",
}

def build_excel(data, output_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "📊 레퍼런스 분석"
    ws.freeze_panes = "D3"
    ws.sheet_view.showGridLines = False

    # 섹션 헤더 (Row 1)
    ws.row_dimensions[1].height = 22
    sec_spans = {}
    for ci, (*_, sec) in enumerate(COLS, 1):
        sec_spans.setdefault(sec, [ci, ci])
        sec_spans[sec][1] = ci
    for sec, (start, end) in sec_spans.items():
        ws.merge_cells(start_row=1, start_column=start, end_row=1, end_column=end)
        c = ws.cell(row=1, column=start)
        c.value = SEC_LABEL[sec]
        c.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
        c.fill = _fill(C[sec])
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = _bdr()

    # 컬럼 헤더 (Row 2)
    ws.row_dimensions[2].height = 36
    for ci, (hdr, _, w, _, sec) in enumerate(COLS, 1):
        c = ws.cell(row=2, column=ci, value=hdr)
        c.font = Font(name="Arial", bold=True, color="FFFFFF", size=9)
        c.fill = _fill(C[sec])
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _bdr()
        ws.column_dimensions[get_column_letter(ci)].width = w

    # 데이터 행
    for ri, game in enumerate(data, 1):
        er = ri + 2
        even = ri % 2 == 0
        ws.row_dimensions[er].height = 82

        for ci, (_, field, _, align, sec) in enumerate(COLS, 1):
            cell = ws.cell(row=er, column=ci)
            is_ai = sec == "s5"
            # Follower/Wishlist는 자동 수집으로 표시 (노란색 제거)
            cell.fill = _fill(C["ai_b"] if is_ai and even else
                              C["ai_a"] if is_ai else
                              C["row_b"] if even else C["row_a"])

            if field is None:
                cell.value = ri
                cell.font = Font(name="Arial", bold=True, size=11, color=C["green"])
                cell.alignment = Alignment(horizontal="center", vertical="top")
            elif field == "name":
                cell.value = game.get(field, "")
                cell.font = Font(name="Arial", bold=True, size=10)
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            elif field == "steam_url":
                url = game.get(field, "")
                cell.value = url; cell.hyperlink = url
                cell.font = Font(name="Arial", size=9, color="0563C1", underline="single")
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            elif field in ("total_reviews","estimated_rev"):
                cell.value = game.get(field, "")
                cell.font = Font(name="Arial", bold=True, size=9, color="1A4A2A")
                cell.alignment = Alignment(horizontal="center", vertical="top")
            elif field in ("followers","wishlist"):
                cell.value = game.get(field, "")
                cell.font = Font(name="Arial", bold=True, size=9, color="0F6E56")
                cell.alignment = Alignment(horizontal="center", vertical="top")
            elif is_ai:
                cell.value = game.get(field, "")
                cell.font = Font(name="Arial", size=9, color="3A1A4A")
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            else:
                cell.value = game.get(field, "")
                cell.font = Font(name="Arial", size=9)
                cell.alignment = Alignment(horizontal=align, vertical="top", wrap_text=True)
            cell.border = _bdr()

    wb.save(output_path)
    print(f"✅ Excel 저장: {output_path}")

# ══════════════════════════════════════════════════════════
#  ⑥ 메인
# ══════════════════════════════════════════════════════════

def main():
    print("=" * 62)
    print("  Steam Competitor Analyzer v3  |  SteamDB Extension 크롤링")
    print("=" * 62)

    # Step 1: Steam API + SteamSpy (빠른 배치 처리)
    print("\n📡 Step 1: Steam API / SteamSpy 데이터 수집")
    all_data = []
    for i, appid in enumerate(APP_IDS, 1):
        print(f"  [{i}/{len(APP_IDS)}] appid={appid}")
        game = collect_steam_data(appid)
        all_data.append(game)
        print(f"    ✓ {game['name']}")

    # Step 2: Playwright — Followers / Wishlist
    print("\n🌐 Step 2: Playwright로 Followers / Wishlist 수집")
    try:
        all_data = run_playwright_scraping(all_data)
    except Exception as e:
        print(f"  [Playwright 전체 실패] {e}")
        print("  Followers/Wishlist를 'N/A'로 처리합니다.")
        for g in all_data:
            if g.get("followers") == "수집 중...":
                g["followers"] = "N/A"
            if g.get("wishlist") == "수집 중...":
                g["wishlist"] = "N/A"

    # Step 3: Claude AI 메타 분석
    print("\n🤖 Step 3: Claude AI 메타 분석")
    for i, game in enumerate(all_data, 1):
        print(f"  [{i}/{len(all_data)}] {game['name']} 분석 중...")
        ai = analyze_with_claude(game)
        game.update(ai)
        print(f"    ✓ {game['core_genre']}")
        time.sleep(1.0)

    # Step 4: Excel 빌드
    print(f"\n📊 Step 4: Excel 빌드 → {OUTPUT_PATH}")
    build_excel(all_data, OUTPUT_PATH)

    print("\n📋 완료:")
    print(f"  게임 수: {len(all_data)}개  |  출력: {OUTPUT_PATH}")
    print("=" * 62)

if __name__ == "__main__":
    main()
